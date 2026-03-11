from __future__ import annotations

from typing import List, Optional
from pathlib import Path
import logging
import time
from urllib.parse import urlparse, parse_qs

from playwright.sync_api import sync_playwright, Page, BrowserContext

from .models import IDUConfig, IDUResult, PEPEntry
from . import session as session_mod
from . import parser as parser_mod

logger = logging.getLogger(__name__)


class IDUScraper:
    """Playwright scraper for idu.tracesmart.co.uk

    Session persistence: manual MFA once, then cookies saved.
    """

    def __init__(
        self,
        username: str,
        password: str,
        session_file: str = "output/sessions/idu_session.json",
        headless: bool = False,
        output_dir: str = "output",
        retry_limit: int = 3,
        slow_mo_ms: int = 0,
    ) -> None:
        self.username = username
        self.password = password
        self.session_file = session_file
        self.output_dir = Path(output_dir)
        self.retry_limit = retry_limit
        self.slow_mo_ms = slow_mo_ms

        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=False, slow_mo=self.slow_mo_ms)
        self.context: BrowserContext = self.browser.new_context(
            viewport={"width": 1280, "height": 800},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
                        " AppleWebKit/537.36 (KHTML, like Gecko)"
                        " Chrome/114.0.0.0 Safari/537.36"),
        )
        self.page: Page = self.context.new_page()

    def _ensure_logged_in(self) -> None:
        """Ensure session is logged in, or perform manual login + MFA."""
        try:
            loaded = session_mod.load_session(self.context, self.session_file)
            if loaded:
                if session_mod.is_session_valid(self.page):
                    logger.info("Session restored, skipping login")
                    return
            # not valid, perform login
            self.page.goto("https://sso.tracesmart.co.uk/login/idu", timeout=20000)
            self.page.wait_for_selector("#username", timeout=20000)
            self.page.fill("#username", self.username)
            self.page.fill("#password", self.password)
            self.page.click('input[data-testid="sign-in"]')
            print("ACTION REQUIRED: Check your email for the MFA code. Enter it in the browser, submit it, wait until you see the IDU dashboard, then press Enter here...")
            input()
            if session_mod.is_session_valid(self.page):
                session_mod.save_session(self.context, self.session_file)
                logger.info("Session saved")
                return
            raise RuntimeError("Login failed after MFA")
        except Exception:
            logger.exception("Error during login flow")
            raise

    def search(self, config: IDUConfig, screenshot: bool = False) -> IDUResult:
        """Run a single search for the provided config."""
        last_exc: Optional[Exception] = None
        for attempt in range(1, self.retry_limit + 1):
            try:
                self._ensure_logged_in()
                self.page.goto("https://idu.tracesmart.co.uk/?page=newSearch&searchtype=1", timeout=20000)
                self.page.wait_for_selector("#forename", timeout=20000)

                # fill fields
                self.page.fill("#forename", config.forename or "")
                if config.middlename:
                    self.page.fill("#middlename", config.middlename)
                if config.surname:
                    self.page.fill("#surname", config.surname)
                if config.reference:
                    self.page.fill("#reference", config.reference)
                self.page.fill("#dd", config.dd or "")
                self.page.fill("#mm", config.mm or "")
                self.page.fill("#yyyy", config.yyyy or "")
                if config.gender:
                    try:
                        self.page.select_option("#gender", config.gender)
                    except Exception:
                        # try by value
                        self.page.select_option("#gender", value=config.gender)
                # address fields
                if config.house:
                    self.page.fill("#house", config.house)
                if config.street:
                    self.page.fill("#street", config.street)
                if config.town:
                    self.page.fill("#town", config.town)
                if config.postcode:
                    self.page.fill("#postcode", config.postcode)
                # contacts
                if config.email:
                    self.page.fill("#email", config.email)
                if config.email2:
                    self.page.fill("#email2", config.email2)
                if config.mobile:
                    self.page.fill("#mobile", config.mobile)
                if config.mobile2:
                    self.page.fill("#mobile2", config.mobile2)
                if config.landline:
                    self.page.fill("#landline", config.landline)
                if config.landline2:
                    self.page.fill("#landline2", config.landline2)

                self.page.click("#addchk")
                # wait for addressmatch link or text
                try:
                    self.page.wait_for_selector("#addressmatch a", timeout=10000)
                    self.page.click("#addressmatch a")
                except Exception:
                    # if no link, try confirming available text
                    try:
                        self.page.wait_for_selector("#addressmatch", timeout=10000)
                    except Exception:
                        logger.info("No addressmatch found; continuing")

                try:
                    self.page.check("#confirm-yes")
                except Exception:
                    # try clicking if not checkable
                    try:
                        self.page.click("#confirm-yes")
                    except Exception:
                        pass

                self.page.click("#inputbut")
                self.page.wait_for_load_state("networkidle", timeout=30000)
                self.page.wait_for_selector("#result-summary-status", timeout=20000)

                screenshot_path = None
                if screenshot:
                    self.output_dir.mkdir(parents=True, exist_ok=True)
                    ts = time.strftime("%Y%m%d_%H%M%S")
                    screenshot_path = str(self.output_dir / f"idu_{config.reference or config.forename}_{ts}.png")
                    try:
                        self.page.screenshot(path=screenshot_path, full_page=True)
                    except Exception:
                        logger.exception("Failed to save screenshot")

                html = self.page.content()
                soup = parser_mod.BeautifulSoup(html, "lxml")

                verdict, score = parser_mod.parse_verdict(soup)
                summary_items = parser_mod.parse_summary_table(soup)
                address_detail = parser_mod.parse_address_section(soup)
                credit_active = parser_mod.parse_credit_active(soup)
                dob_ver = parser_mod.parse_dob_verification(soup)
                pep_entries, sanction_result = parser_mod.parse_pep_sanctions(soup)
                mortality = parser_mod.parse_section_by_id(soup, "res-mortality-body")
                gone_away = parser_mod.parse_section_by_id(soup, "res-goneaway-body")
                ccj = parser_mod.parse_section_by_id(soup, "res-ccj-body")
                insolvency = parser_mod.parse_section_by_id(soup, "res-insolvency-body")
                company_director = parser_mod.parse_section_by_id(soup, "res-companydirector-body")
                search_activity = parser_mod.parse_section_by_id(soup, "res-searchactivity-body")
                address_links = parser_mod.parse_address_links(soup)
                property_detail = parser_mod.parse_property(soup)

                # extract search id
                parsed = urlparse(self.page.url)
                qs = parse_qs(parsed.query)
                search_id = qs.get("id", [""])[0]

                result = IDUResult(
                    config = (config.__dict__ if hasattr(config, "__dict__") else {}),
                    scraped_at = time.strftime("%Y-%m-%d %H:%M:%S"),
                    search_id = search_id,
                    verdict = verdict,
                    score = score,
                    summary_items = summary_items,
                    address_detail = address_detail,
                    credit_active = credit_active,
                    dob_verification = dob_ver,
                    mortality = mortality,
                    gone_away = gone_away,
                    pep_entries = pep_entries,
                    sanction_result = sanction_result,
                    ccj = ccj,
                    insolvency = insolvency,
                    company_director = company_director,
                    search_activity = search_activity,
                    address_links = address_links,
                    property_detail = property_detail,
                    screenshot_path = screenshot_path,
                )

                return result

            except Exception as exc:  # per-attempt
                logger.warning("Search attempt %s failed: %s", attempt, exc)
                last_exc = exc
                if attempt < self.retry_limit:
                    backoff = 2 ** attempt
                    time.sleep(backoff)
                    continue
                else:
                    # return result with error
                    res = IDUResult(config=(config.__dict__ if hasattr(config, "__dict__") else {}), scraped_at=time.strftime("%Y-%m-%d %H:%M:%S"))
                    res.error = str(last_exc)
                    return res

        # unreachable
        res = IDUResult(config=(config.__dict__ if hasattr(config, "__dict__") else {}), scraped_at=time.strftime("%Y-%m-%d %H:%M:%S"))
        res.error = "Exceeded retries"
        return res

    def search_batch(self, configs: List[IDUConfig], screenshot: bool = False, save_xlsx: bool = False):
        """Process batch of configs and optionally save XLSX."""
        results = []
        total = len(configs)
        for i, cfg in enumerate(configs):
            logger.info("Processing %d/%d: %s %s", i+1, total, cfg.forename, cfg.surname or "")
            res = self.search(cfg, screenshot=screenshot)
            results.append(res)
            time.sleep(2)

        if save_xlsx:
            try:
                path = self.save_results_xlsx(results)
                logger.info("Saved XLSX to %s", path)
            except Exception:
                logger.exception("Failed to save XLSX")

        return results

    def save_results_xlsx(self, results: List[IDUResult], filename: Optional[str] = None):
        """Save results to an Excel file with Summary and PEP Detail sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws2 = wb.create_sheet("PEP Detail")

        headers = [
            "Reference","Forename","Surname","DOB","Postcode",
            "Verdict","Score","Date of Search",
            "Electoral Roll","Tracesmart Register","Credit Active",
            "DOB Verification","PEP Alert","Mortality","Gone Away",
            "CCJ","Insolvency","Company Director",
            "PEP Count","Sanction Result","Error",
        ]
        ws1.append(headers)
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        thin = Side(border_style="thin", color="000000")
        for col in range(1, len(headers)+1):
            cell = ws1.cell(row=1, column=col)
            cell.font = bold
            cell.fill = header_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        pep_headers = [
            "Reference","Forename","Surname","DOB",
            "Match Score","PEP Name","Aliases","Last Updated",
            "Addresses","Country","Position","Reason",
        ]
        ws2.append(pep_headers)
        for cell in ws2[1]:
            cell.font = bold
            cell.fill = header_fill

        # populate
        for r in results:
            cfg = r.config or {}
            # find summary item statuses by label
            def lookup(label):
                for it in r.summary_items:
                    if it.label.lower() == label.lower():
                        return it.status
                return "not_checked"

            pep_count = len(r.pep_entries or [])
            row = [
                cfg.get("reference",""), cfg.get("forename",""), cfg.get("surname",""),
                f"{cfg.get('dd','')}/{cfg.get('mm','')}/{cfg.get('yyyy','')}", cfg.get("postcode",""),
                r.verdict, r.score, r.date_of_search,
                lookup("Electoral Roll"), lookup("Tracesmart Register"), lookup("Credit Active"),
                lookup("DOB Verification"), ("alert" if pep_count>0 else "pass"),
                lookup("Mortality"), lookup("Gone Away"),
                lookup("CCJ"), lookup("Insolvency"), lookup("Company Director"),
                pep_count, r.sanction_result, r.error or "",
            ]
            ws1.append(row)

            for pep in (r.pep_entries or []):
                prow = [
                    cfg.get("reference",""), cfg.get("forename",""), cfg.get("surname",""),
                    f"{cfg.get('dd','')}/{cfg.get('mm','')}/{cfg.get('yyyy','')}",
                    pep.match_score, pep.name, ", ".join(pep.aliases), pep.last_updated,
                    ", ".join(pep.addresses), pep.country, pep.position, pep.reason,
                ]
                ws2.append(prow)

        # autofit columns (naive)
        for ws in (ws1, ws2):
            for col in ws.columns:
                maxlen = 0
                for cell in col:
                    try:
                        val = str(cell.value or "")
                        if len(val) > maxlen:
                            maxlen = len(val)
                    except Exception:
                        pass
                width = min(50, max(10, maxlen + 2))
                ws.column_dimensions[col[0].column_letter].width = width
            ws.freeze_panes = "A2"

        if not filename:
            ts = time.strftime("%Y%m%d_%H%M%S")
            filename = str(self.output_dir / f"idu_results_{ts}.xlsx")
        else:
            filename = str(self.output_dir / filename)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(filename)
        return filename

    def close(self) -> None:
        try:
            self.page.close()
        except Exception:
            pass
        try:
            self.context.close()
        except Exception:
            pass
        try:
            self.browser.close()
        except Exception:
            pass
        try:
            self.playwright.stop()
        except Exception:
            pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()
